import os

from fastapi import APIRouter
from fastapi.responses import FileResponse
from pydantic import BaseModel
from extract.extract_table import get_time_table
import json
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from io import BytesIO

from api.config.redis_config import (
    get_table_from_cache,
    add_table_to_cache,
)


# Find the path of the drafts
current_script_path = Path(__file__)
project_root_path = current_script_path.parents[1]
DRAFTS_FOLDER = project_root_path / "drafts"

router = APIRouter()


class TimeTableRequest(BaseModel):
    """
    Represents a request for a timetable.

    Attributes:
    - filename (str): The name of the file for the timetable.
    - class_pattern (str): The pattern for the class.
    """

    filename: str
    class_pattern: str

def get_json_table(request: TimeTableRequest):
    """
    A function to get the time table in JSON format.

    Parameters:
    - request: TimeTableRequest - the request object containing the filename and class pattern

    Returns:
    - dict: a dictionary containing the table in JSON format
    """
    filename = os.path.join(DRAFTS_FOLDER, request.filename)

    table = get_table_from_cache(request.class_pattern, request.filename)

    if table is None:
        table = get_time_table(filename, request.class_pattern).to_json(
            orient="records"
        )
        add_table_to_cache(
            table=table, class_pattern=request.class_pattern, filename=request.filename
        )

    return json.loads(table)

@router.post("/get_time_table")
async def get_time_table(request: TimeTableRequest):
    """
    Endpoint for generating a parsed json time table

    Args:
        request (TimeTableRequest): The request object containing the filename and class pattern.

    Returns:
        JSON: Parsed data from the get_json_table function that contains the time table cutting across days and time slots. 
        It covers merged durations of lectures exceeding one hour as well.
    """    
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    json_data = get_json_table(request)

    table_data = []
    for index, day in enumerate(json_data):
        day_data = []
        current_slot = None
        for key, value in day.items():
            time_parts = key.split("-")
            if len(time_parts) == 2:
                start, end = time_parts
            else:
                start = "-".join(time_parts[:-1])
                end = time_parts[-1]
            if current_slot and current_slot["value"] == value and current_slot["end"] == start:
                current_slot["end"] = end
            else:
                if current_slot:
                    day_data.append(current_slot)
                current_slot = {"start": start, "end": end, "value": value}
        if current_slot:
            day_data.append(current_slot)
        table_data.append({"day": days[index], "data": day_data})

    return table_data


@router.post("/download")
async def download_time_table_endpoint(request: TimeTableRequest):
    """
    Endpoint for downloading a time table as an Excel file.

    Parameters:
    - request (TimeTableRequest): The request object containing the filename and class pattern.

    Returns:
    - FileResponse: The Excel file containing the time table.

    Description:
    This function is an endpoint for downloading a time table as an Excel file. 
    It takes a `TimeTableRequest` object as a parameter, which contains the filename and class pattern.
    The function first checks if the time table is already cached. If it is, it retrieves the cached table.
    Otherwise, it generates the time table by calling the `get_time_table` function and adds it to the cache.
    The function then converts the time table into a Pandas DataFrame and creates an Excel file using the `openpyxl` library.
    It iterates over the columns and rows of the DataFrame and writes the values to the Excel worksheet. 
    Finally, it saves the Excel file to a buffer and returns it as a `FileResponse` object with the appropriate media type.

    Note:
    - The `TimeTableRequest` class should have the following attributes:
        - filename (str): The name of the file for the time table.
        - class_pattern (str): The pattern for the class.
    - The `get_table_from_cache` and `add_table_to_cache` functions should be implemented elsewhere in the codebase.

    ```
    """
    filename = os.path.join(DRAFTS_FOLDER, request.filename)

    table = get_table_from_cache(request.class_pattern, request.filename)

    if table is None:
        table = get_time_table(filename, request.class_pattern).to_json(
            orient="records"
        )
        add_table_to_cache(
            table=table, class_pattern=request.class_pattern, filename=request.filename
        )

        df = pd.DataFrame(table)
        buffer = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active

        for col_index, col_name in enumerate(df.columns, start=1):
            worksheet.cell(row=1, column=col_index, value=col_name)

        for row_index, row in enumerate(df.itertuples(), start=2):
            for col_index, value in enumerate(row[1:], start=1):
                worksheet.cell(row=row_index, column=col_index, value=value)

        workbook.save(buffer)

        excel_content = buffer.getvalue()
        return FileResponse(
            excel_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
