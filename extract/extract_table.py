import regex as re
import pandas as pd
from icalendar import Event, Calendar
from datetime import datetime, timedelta
import json
import pytz
import openpyxl


def _get_time_row(df: pd.DataFrame) -> pd.Series:
    """
    Get the time row from the dataframe.

    Parameters
    ----------
    df : pandas.DataFrame
        The dataframe to get the time row from.

    Returns
    -------
    pandas.Series
        The time row from the dataframe.
    """
    for row in df.iterrows():
        if re.match(r"^\d{1,2}:\d{1,2}-\d{1,2}:\d{1,2}$", str(row[1].iloc[1])):
            return row


def _get_daily_table(df: pd.DataFrame, class_pattern: str) -> pd.DataFrame:
    """
        Get the a simplified dataframe of the classes for a given class.

        Parameters
        ----------    table = catched_get_table(raw_file, class_to_extract_for)

        df : pandas.DataFrame
            The dataframe to get the simplified time table from.
            It's a general time table on a single day for all classes.
        class_pattern : str
            The class to search for. E.g. 'EL 3'
    o
        Returns
        -------
        pandas.DataFrame
            The simplified dataframe for only the given class.
    """
    df = df.copy()

    time_row = _get_time_row(df)
    new_cols = time_row[1].to_list()
    new_cols.pop(0)
    new_cols.insert(0, "Classroom")
    df.columns = new_cols

    df.set_index("Classroom", inplace=True)

    df = df.iloc[time_row[0] + 1 :]

    df = df.mask(~df.map(lambda x: bool(re.search(class_pattern, str(x)))))
    df = df.dropna(how="all")

    return df


def _get_all_daily_tables(filename: str, class_pattern: str) -> dict:
    """
    Get all the daily tables from an excel file.

    Parameters
    ----------
    filename : str
        The filename of the excel file to get the daily tables from.
    class_pattern : str
        The class to get the daily tables or. E.g. 'EL 3'

    Returns
    -------
    dict
        A dictionary of the daily tables for each class.
    """
    filename += ".xlsx"

    workbook = openpyxl.load_workbook(filename)
    dfs = {}
    for sheet in workbook.sheetnames:
        merged_cells = workbook[sheet].merged_cells.ranges
        for mc in merged_cells.copy():
            if mc.max_col - mc.min_col == 1:
                merged_value = workbook[sheet].cell(mc.min_row, mc.min_col).value
                workbook[sheet].unmerge_cells(mc.coord)
                workbook[sheet].cell(mc.min_row, mc.min_col).value = merged_value
                workbook[sheet].cell(mc.max_row, mc.max_col).value = merged_value

        data = workbook[sheet].values
        header = next(data)
        df = pd.DataFrame(data, columns=header)
        df = df.dropna(axis=1, how="all")

        dfs[sheet] = _get_daily_table(df, class_pattern)

    return dfs


def get_time_table2(filename: str, class_pattern: str) -> pd.DataFrame:
    """
    Get the complete time table for a particular class for all days.

    Parameters
    ----------
    filename : str
        The filename of the excel file. This file contains every class with the days as the sheet names.
    class_pattern : str
        The class to get the complete time table for. E.g. 'EL 3'

    Returns
    -------
    pandas.DataFrame
        The complete time table for the given class.
    """
    daily_tables = _get_all_daily_tables(filename, class_pattern)

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    for key, value in daily_tables.items():
        if key.title() in days:
            columns = value.columns
            break
    else:
        raise ValueError(f"No sheet found for any of the days: {days}")

    final_df = pd.DataFrame(
        columns=columns,
        index=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
    )

    for day, table in daily_tables.items():
        for period, classes in table.items():
            available_classes = classes.dropna()
            if available_classes.any():
                classrooms = classes[classes.notna()].index
                available_classes = [
                    re.sub(r"\s+", " ", c.strip()) for c in available_classes.values
                ]
                available_classes = [
                    f"{c} ({classrooms[i]})" for i, c in enumerate(available_classes)
                ]
                available_classes = "\n".join(available_classes)
                final_df.loc[day, period] = available_classes

    return final_df


def convert_to_datetime(obj):
    if isinstance(obj, datetime):
        return obj
    elif isinstance(obj, datetime.date):
        return datetime(obj.year, obj.month, obj.day)
    elif isinstance(obj, int):
        # Example conversion: if obj is a timestamp
        return datetime.fromtimestamp(obj)
    else:
        raise TypeError("Unsupported type for datetime conversion")


def generate_calendar(timetable, start_date, end_date):
    """
    Generate a calendar of class events based on a given timetable within a specified date range.

    Parameters:
        timetable (list): A list of dictionaries representing the timetable data. Each dictionary contains the following keys:
            - day (str): The name of the day.
            - data (list): A list of dictionaries representing the class events for the day. Each dictionary contains the following keys:
                - start (str): The start time of the class in the format 'HH:MM'.
                - end (str): The end time of the class in the format 'HH:MM'.
                - value (str): The name of the class.

        start_date (str): The start date of the calendar in the format 'YYYY-MM-DD'.
        end_date (str): The end date of the calendar in the format 'YYYY-MM-DD'.

    Returns:
        None

    This function generates a calendar of class events based on the provided timetable within the specified date range.
    It iterates over the timetable data, checks if the day matches the current date, and adds class events to the calendar.
    The resulting calendar is saved as an ICS file named 'class_schedule.ics'.
    """
    data = timetable

    cal = Calendar()

    start_date = datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.strptime(end_date, "%Y-%m-%d")

    current_date = start_date
    while current_date <= end_date:
        day_name = current_date.strftime("%A")
        for day in data:
            if day["day"] == day_name:
                for class_info in day["data"]:
                    if (
                        class_info["start"]
                        and class_info["end"]
                        and class_info["value"]
                    ):
                        # Parse start and end times in 24-hour format
                        start_time = datetime.strptime(class_info["start"], "%H:%M")
                        end_time = datetime.strptime(class_info["end"], "%H:%M")

                        event = Event()
                        event.add("summary", class_info["value"].replace("\n", " "))
                        event.add(
                            "dtstart",
                            current_date.replace(
                                hour=start_time.hour, minute=start_time.minute
                            ),
                        )
                        event.add(
                            "dtend",
                            current_date.replace(
                                hour=end_time.hour, minute=end_time.minute
                            ),
                        )
                        cal.add_component(event)
        current_date += timedelta(days=1)

    with open("class_schedule.ics", "wb") as f:
        f.write(cal.to_ical())
        return cal.to_ical()